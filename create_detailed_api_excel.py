#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_detailed_api_excel():
    wb = Workbook()
    
    # 기본 스타일 정의
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    wrap_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # API 목록 시트 생성
    ws_summary = wb.active
    ws_summary.title = "API_목록"
    
    # API 목록 헤더
    summary_headers = ["번호", "URI", "설명", "주기", "메서드", "인증"]
    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # API 목록 데이터
    api_list = [
        [1, "/", "API 서버 상태 확인", "실시간", "GET", "불필요"],
        [2, "/health", "시스템 전반 상태 확인", "실시간", "GET", "Bearer Token"],
        [3, "/auth", "사용자 인증 및 토큰 발급", "실시간", "POST", "불필요"],
        [4, "/verify-token", "JWT 토큰 유효성 검증", "실시간", "POST", "불필요"],
        [5, "/companies", "등록된 회사 정보 목록 조회", "실시간", "GET", "Bearer Token"],
        [6, "/products", "등록된 제품 정보 목록 조회", "실시간", "GET", "Bearer Token"],
        [7, "/clients", "등록된 병원 정보 목록 조회", "실시간", "GET", "Bearer Token"],
        [8, "/pharmacies", "등록된 약국 정보 목록 조회", "실시간", "GET", "Bearer Token"],
        [9, "/notices", "시스템 공지사항 목록 조회", "실시간", "GET", "Bearer Token"],
        [10, "/hospital-company-mappings", "병원과 업체 간의 관계 정보 조회", "실시간", "GET", "Bearer Token"],
        [11, "/hospital-pharmacy-mappings", "병원과 약국 간의 관계 정보 조회", "실시간", "GET", "Bearer Token"],
        [12, "/client-company-assignments", "병원과 업체 간의 배정 정보 조회", "실시간", "GET", "Bearer Token"],
        [13, "/client-pharmacy-assignments", "병원과 약국 간의 배정 정보 조회", "실시간", "GET", "Bearer Token"],
        [14, "/product-company-not-assignments", "업체에 배정되지 않은 제품 정보 조회", "실시간", "GET", "Bearer Token"],
        [15, "/wholesale-sales", "도매 매출 데이터 조회", "실시간", "GET", "Bearer Token"],
        [16, "/direct-sales", "직매 매출 데이터 조회", "실시간", "GET", "Bearer Token"],
        [17, "/performance-records", "실적 정보 목록 조회", "실시간", "GET", "Bearer Token"],
        [18, "/performance-records-absorption", "실적 흡수율 분석 정보 조회", "실시간", "GET", "Bearer Token"],
        [19, "/performance-evidence-files", "실적 증빙 파일 정보 조회", "실시간", "GET", "Bearer Token"],
        [20, "/settlement-months", "정산월 목록 조회", "실시간", "GET", "Bearer Token"],
        [21, "/settlement-share", "정산내역서 목록 조회", "실시간", "GET", "Bearer Token"]
    ]
    
    for row, data in enumerate(api_list, 2):
        for col, value in enumerate(data, 1):
            cell = ws_summary.cell(row=row, column=col, value=value)
            cell.border = border
            if col in [1, 2, 5, 6]:
                cell.alignment = center_alignment
            else:
                cell.alignment = wrap_alignment
    
    # 컬럼 너비 조정
    ws_summary.column_dimensions['A'].width = 8
    ws_summary.column_dimensions['B'].width = 35
    ws_summary.column_dimensions['C'].width = 40
    ws_summary.column_dimensions['D'].width = 12
    ws_summary.column_dimensions['E'].width = 10
    ws_summary.column_dimensions['F'].width = 15
    
    # 각 API별 상세 시트 생성
    api_details = [
        {
            "name": "API_상태확인",
            "uri": "/",
            "description": "API 서버 상태 확인",
            "frequency": "실시간",
            "method": "GET",
            "auth": "불필요",
            "input_params": [],
            "output_params": [
                {"name": "name", "type": "STRING", "description": "API 서버명", "required": "Y", "remarks": ""},
                {"name": "version", "type": "STRING", "description": "API 버전", "required": "Y", "remarks": ""},
                {"name": "status", "type": "STRING", "description": "서버 상태", "required": "Y", "remarks": ""},
                {"name": "timestamp", "type": "STRING", "description": "응답 시간", "required": "Y", "remarks": ""},
                {"name": "environment", "type": "STRING", "description": "환경 정보", "required": "Y", "remarks": ""}
            ]
        },
        {
            "name": "시스템_헬스체크",
            "uri": "/health",
            "description": "시스템 전반 상태 확인",
            "frequency": "실시간",
            "method": "GET",
            "auth": "Bearer Token",
            "input_params": [],
            "output_params": [
                {"name": "status", "type": "STRING", "description": "시스템 상태", "required": "Y", "remarks": ""},
                {"name": "timestamp", "type": "STRING", "description": "응답 시간", "required": "Y", "remarks": ""},
                {"name": "uptime", "type": "FLOAT", "description": "서버 가동시간(초)", "required": "Y", "remarks": ""},
                {"name": "environment", "type": "STRING", "description": "환경 정보", "required": "Y", "remarks": ""}
            ]
        },
        {
            "name": "사용자_로그인",
            "uri": "/auth",
            "description": "사용자 인증 및 토큰 발급",
            "frequency": "실시간",
            "method": "POST",
            "auth": "불필요",
            "input_params": [
                {"name": "email", "type": "STRING", "description": "사용자 이메일", "required": "Y", "remarks": ""},
                {"name": "password", "type": "STRING", "description": "사용자 비밀번호", "required": "Y", "remarks": ""}
            ],
            "output_params": [
                {"name": "success", "type": "BOOLEAN", "description": "성공 여부", "required": "Y", "remarks": ""},
                {"name": "token", "type": "STRING", "description": "JWT 토큰", "required": "Y", "remarks": ""},
                {"name": "user", "type": "OBJECT", "description": "사용자 정보", "required": "Y", "remarks": ""},
                {"name": "message", "type": "STRING", "description": "응답 메시지", "required": "Y", "remarks": ""}
            ]
        }
    ]
    
    # 상세 시트 생성 (처음 3개 API만 예시로 생성)
    for api in api_details:
        ws = wb.create_sheet(api["name"])
        
        # API 기본 정보 테이블
        current_row = 1
        
        # API 정보 헤더
        info_headers = ["항목", "내용"]
        for col, header in enumerate(info_headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        current_row += 1
        
        # API 정보 데이터
        api_info = [
            ["URI", api["uri"]],
            ["설명", api["description"]],
            ["주기", api["frequency"]],
            ["메서드", api["method"]],
            ["인증", api["auth"]]
        ]
        
        for info in api_info:
            for col, value in enumerate(info, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.border = border
                if col == 1:
                    cell.font = Font(bold=True)
                    cell.alignment = center_alignment
                else:
                    cell.alignment = wrap_alignment
            current_row += 1
        
        current_row += 2
        
        # 입력 파라미터 테이블
        if api["input_params"]:
            ws.cell(row=current_row, column=1, value="입력 파라미터").font = Font(bold=True, size=12)
            current_row += 1
            
            input_headers = ["파라미터명", "타입", "설명", "필수"]
            for col, header in enumerate(input_headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border
            
            current_row += 1
            
            for param in api["input_params"]:
                for col, value in enumerate([param["name"], param["type"], param["description"], param["required"]], 1):
                    cell = ws.cell(row=current_row, column=col, value=value)
                    cell.border = border
                    cell.alignment = wrap_alignment
                current_row += 1
        else:
            ws.cell(row=current_row, column=1, value="입력 파라미터").font = Font(bold=True, size=12)
            current_row += 1
            ws.cell(row=current_row, column=1, value="없음").border = border
            current_row += 1
        
        current_row += 2
        
        # 출력 파라미터 테이블
        ws.cell(row=current_row, column=1, value="출력 파라미터").font = Font(bold=True, size=12)
        current_row += 1
        
        output_headers = ["파라미터명", "타입", "설명", "필수", "비고"]
        for col, header in enumerate(output_headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        current_row += 1
        
        for param in api["output_params"]:
            for col, value in enumerate([param["name"], param["type"], param["description"], param["required"], param["remarks"]], 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.border = border
                cell.alignment = wrap_alignment
            current_row += 1
        
        # 컬럼 너비 조정
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 15
    
    # 파일 저장
    filename = "Sinil_PMS_API_Detailed.xlsx"
    wb.save(filename)
    
    print(f"✅ 상세 Excel 파일이 생성되었습니다: {filename}")
    print(f"📊 API 목록 시트와 상세 파라미터 시트가 포함되었습니다.")
    
    return filename

if __name__ == "__main__":
    create_detailed_api_excel()
