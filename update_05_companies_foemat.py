import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

def update_companies_excel():
    # Excel 파일 로드
    file_path = 'API_Files/05_회사정보_조회.xlsx'
    wb = load_workbook(file_path)
    ws = wb.active
    
    # 새로운 필드 설명 추가
    new_row = 15
    
    # 제목 스타일
    title_font = Font(bold=True, size=12)
    title_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    
    # 새로운 필드 설명
    ws.cell(row=new_row, column=1, value="추가 필드 설명")
    ws.cell(row=new_row, column=1).font = title_font
    ws.cell(row=new_row, column=1).fill = title_fill
    
    ws.cell(row=new_row+1, column=1, value="change_type")
    ws.cell(row=new_row+1, column=2, value="변경 유형")
    ws.cell(row=new_row+1, column=3, value="'new': 신규 생성, 'updated': 수정됨")
    
    ws.cell(row=new_row+2, column=1, value="last_modified")
    ws.cell(row=new_row+2, column=2, value="최근 변경 시간")
    ws.cell(row=new_row+2, column=3, value="updated_at이 있으면 updated_at, 없으면 created_at")
    
    # 기존 설명도 추가
    ws.cell(row=new_row+4, column=1, value="날짜 검색 조건 상세 설명")
    ws.cell(row=new_row+4, column=1).font = title_font
    ws.cell(row=new_row+4, column=1).fill = title_fill
    
    ws.cell(row=new_row+5, column=1, value="startDate")
    ws.cell(row=new_row+5, column=2, value="시작 날짜 (YYYY-MM-DD)")
    ws.cell(row=new_row+5, column=3, value="created_at과 updated_at 모두 검색")
    
    ws.cell(row=new_row+6, column=1, value="endDate")
    ws.cell(row=new_row+6, column=2, value="종료 날짜 (YYYY-MM-DD)")
    ws.cell(row=new_row+6, column=3, value="created_at과 updated_at 모두 검색")
    
    ws.cell(row=new_row+7, column=1, value="중복 제거")
    ws.cell(row=new_row+7, column=2, value="DISTINCT 적용")
    ws.cell(row=new_row+7, column=3, value="같은 회사가 두 조건에 모두 해당해도 한 번만 반환")
    
    # 파일 저장
    wb.save(file_path)
    print(f"✅ {file_path} 파일이 성공적으로 업데이트되었습니다!")

if __name__ == "__main__":
    update_companies_excel()