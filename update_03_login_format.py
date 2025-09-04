#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
03_사용자_로그인.xlsx 파일의 출력 파라미터를 새로운 형식에 맞춰 수정
"""

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import os

def update_03_login_format():
    """03_사용자_로그인.xlsx 파일을 새로운 응답 형식에 맞춰 수정"""
    
    input_file = "API_Files/03_사용자_로그인.xlsx"
    output_file = "API_Files/03_사용자_로그인_updated.xlsx"
    
    # 기존 파일 읽기
    try:
        df = pd.read_excel(input_file)
        print("기존 파일 읽기 성공")
        print(df.to_string())
    except Exception as e:
        print(f"파일 읽기 실패: {e}")
        return
    
    # 새로운 출력 파라미터 정의 (이미지 기반)
    new_output_params = [
        ['파라미터명', '타입', '설명', '필수', '비고'],
        ['success', 'BOOLEAN', '성공 여부', 'Y', 'true/false'],
        ['message', 'STRING', '응답 메시지', 'Y', '인증 성공/실패 메시지'],
        ['data.token', 'STRING', 'JWT 토큰', 'Y', '인증 토큰'],
        ['data.refreshToken', 'STRING', '리프레시 토큰', 'Y', '토큰 갱신용'],
        ['data.user', 'OBJECT', '사용자 정보', 'Y', '로그인한 사용자 정보'],
        ['data.expiresIn', 'STRING', '토큰 만료 시간', 'Y', '예: 24h'],
        ['data.expiresAt', 'STRING', '토큰 만료 일시', 'Y', 'ISO 8601 형식']
    ]
    
    # 새로운 워크북 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "03_사용자_로그인"
    
    # 헤더 스타일 설정
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 기본 정보 (URI, 설명, 주기, 메서드, 인증)
    basic_info = [
        ['항목', '내용', '', '', ''],
        ['URI', '/auth', '', '', ''],
        ['설명', '사용자 인증 및 토큰 발급', '', '', ''],
        ['주기', '실시간', '', '', ''],
        ['메서드', 'POST', '', '', ''],
        ['인증', '불필요', '', '', ''],
        ['', '', '', '', ''],
        ['', '', '', '', ''],
    ]
    
    # 입력 파라미터
    input_params = [
        ['입력 파라미터', '', '', '', ''],
        ['파라미터명', '타입', '설명', '필수', ''],
        ['email', 'STRING', '사용자 이메일', 'Y', ''],
        ['password', 'STRING', '사용자 비밀번호', 'Y', ''],
        ['', '', '', '', ''],
        ['', '', '', '', ''],
    ]
    
    # 출력 파라미터
    output_params = [
        ['출력 파라미터', '', '', '', ''],
    ]
    
    # 모든 데이터를 하나의 리스트로 결합
    all_data = basic_info + input_params + output_params + new_output_params[1:]  # 헤더 제외
    
    # 데이터 쓰기
    for row_idx, row_data in enumerate(all_data, 1):
        for col_idx, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            
            # 헤더 행 스타일 적용
            if row_idx == 1 or row_idx in [9, 15]:  # 파라미터 헤더 행들
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
    
    # 열 너비 자동 조정
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 파일 저장
    try:
        wb.save(output_file)
        print(f"\n✅ 파일 수정 완료: {output_file}")
        
        # 수정된 내용 확인
        print("\n=== 수정된 파일 내용 ===")
        updated_df = pd.read_excel(output_file)
        print(updated_df.to_string())
        
    except Exception as e:
        print(f"파일 저장 실패: {e}")
        return
    
    # 원본 파일 교체
    try:
        os.remove(input_file)
        os.rename(output_file, input_file)
        print(f"\n✅ 원본 파일 교체 완료: {input_file}")
    except Exception as e:
        print(f"파일 교체 실패: {e}")

if __name__ == "__main__":
    update_03_login_format()
