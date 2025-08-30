#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

def fix_19_file():
    """19번 파일을 실제 API 응답 구조에 맞게 수정합니다."""
    
    try:
        # 엑셀 파일 로드
        wb = load_workbook('API_Files/19_실적증빙파일.xlsx')
        ws = wb.active
        
        print("=== 19번 파일을 실제 API 응답 구조로 수정 시작 ===")
        
        # 출력 파라미터 섹션 찾기
        output_section_row = None
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and "출력 파라미터" in str(cell_value):
                output_section_row = row
                print(f"출력 파라미터 섹션 발견: 행 {row}")
                break
        
        if output_section_row:
            # 출력 파라미터 섹션의 끝 찾기
            output_end_row = output_section_row
            for row in range(output_section_row + 1, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=1).value
                if cell_value and ("입력 파라미터" in str(cell_value) or "응답 예시" in str(cell_value)):
                    output_end_row = row - 1
                    break
                if row == ws.max_row:
                    output_end_row = row
            
            print(f"출력 파라미터 섹션 끝: 행 {output_end_row}")
            
            # 기존 출력 파라미터 행들 삭제 (헤더 제외)
            for row in range(output_end_row, output_section_row, -1):
                ws.delete_rows(row)
            
            # 새로운 출력 파라미터 정의
            new_output_params = [
                ["success", "BOOLEAN", "성공 여부", "Y", "true/false"],
                ["message", "STRING", "응답 메시지", "Y", "실적증빙파일 조회 결과 메시지"],
                ["data", "ARRAY", "실적증빙파일 목록", "Y", "실적증빙파일 배열 (현재 빈 배열)"],
                ["pagination", "OBJECT", "페이지네이션 정보", "Y", "페이지네이션 관련 정보"]
            ]
            
            # 새로운 행들 삽입
            start_row = output_section_row + 2  # 헤더 다음 행부터
            
            for i, param in enumerate(new_output_params):
                row_num = start_row + i
                
                # 각 셀에 값 설정
                ws.cell(row=row_num, column=1, value=param[0])  # 필드명
                ws.cell(row=row_num, column=2, value=param[1])  # 데이터 타입
                ws.cell(row=row_num, column=3, value=param[2])  # 설명
                ws.cell(row=row_num, column=4, value=param[3])  # 필수 여부
                ws.cell(row=row_num, column=5, value=param[4])  # 상세 설명
                
                # 스타일 적용
                for col in range(1, 6):
                    cell = ws.cell(row=row_num, column=col)
                    cell.font = Font(size=10)
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    cell.alignment = Alignment(horizontal='left', vertical='center')
            
            print(f"새로운 출력 파라미터 {len(new_output_params)}개를 추가했습니다.")
            
            # 파일 저장
            wb.save('API_Files/19_실적증빙파일.xlsx')
            print("19번 파일 수정이 완료되었습니다.")
            
        else:
            print("출력 파라미터 섹션을 찾을 수 없습니다.")
            
    except Exception as e:
        print(f"오류 발생: {str(e)}")

if __name__ == "__main__":
    fix_19_file()
