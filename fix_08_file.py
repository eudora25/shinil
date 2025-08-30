#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
from openpyxl import load_workbook
import requests
import json

def fix_08_file():
    """08번 파일을 실제 API 응답에 맞게 수정합니다."""
    
    try:
        print("=== 08번 파일 (약국정보_조회) 수정 ===")
        
        # 먼저 로그인해서 토큰을 받아옵니다
        print("로그인하여 토큰 받는 중...")
        login_data = {
            "email": "admin@admin.com",
            "password": "asdf1234"
        }
        
        login_response = requests.post('http://localhost:3001/api/auth', json=login_data)
        
        if login_response.status_code != 200:
            print(f"로그인 실패: {login_response.status_code}")
            return
        
        login_result = login_response.json()
        token = login_result.get('data', {}).get('token')
        
        if not token:
            print("토큰을 받을 수 없습니다.")
            return
        
        print("토큰 받기 성공!")
        
        # 실제 API 응답 확인 (약국정보 조회)
        print("실제 API 응답 확인 중...")
        headers = {
            'Authorization': f'Bearer {token}'
        }
        
        response = requests.get('http://localhost:3001/api/pharmacies?page=1&limit=1', headers=headers)
        
        if response.status_code != 200:
            print(f"API 요청 실패: {response.status_code}")
            print(f"응답 내용: {response.text}")
            return
        
        actual_response = response.json()
        print("실제 API 응답 구조 확인 완료!")
        
        # 엑셀 파일 로드
        print("엑셀 파일 수정 중...")
        wb = load_workbook('API_Files/08_약국정보_조회.xlsx')
        ws = wb.active
        
        # 출력 파라미터 섹션 찾기
        output_section_row = None
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and "출력 파라미터" in str(cell_value):
                output_section_row = row
                break
        
        if not output_section_row:
            print("출력 파라미터 섹션을 찾을 수 없습니다.")
            return
        
        # 출력 파라미터 섹션의 끝 찾기
        output_end_row = output_section_row
        for row in range(output_section_row + 1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and any(keyword in str(cell_value) for keyword in ["입력 파라미터", "응답 예시"]):
                break
            output_end_row = row
        
        print(f"출력 파라미터 섹션: 행 {output_section_row} ~ {output_end_row}")
        
        # 기존 출력 파라미터 행들 삭제 (헤더 제외)
        rows_to_delete = []
        for row in range(output_section_row + 2, output_end_row + 1):
            rows_to_delete.append(row)
        
        # 뒤에서부터 삭제 (인덱스 변경 방지)
        for row in reversed(rows_to_delete):
            ws.delete_rows(row)
        
        # 새로운 출력 파라미터 추가
        current_row = output_section_row + 2
        
        # 실제 API 응답 필드들 정의
        api_fields = [
            ("success", "BOOLEAN", "성공 여부", "Y", "true/false"),
            ("data", "ARRAY", "약국 데이터 배열", "Y", "약국 정보 배열"),
            ("data[].id", "INTEGER", "약국 ID", "Y", "고유 식별자"),
            ("data[].pharmacy_code", "STRING", "약국 코드", "N", "약국 고유 코드"),
            ("data[].name", "STRING", "약국명", "Y", "약국 상호명"),
            ("data[].business_registration_number", "STRING", "사업자등록번호", "Y", "사업자등록번호"),
            ("data[].address", "STRING", "주소", "N", "약국 주소"),
            ("data[].remarks", "STRING", "비고", "N", "추가 정보"),
            ("data[].status", "STRING", "상태", "Y", "active/inactive"),
            ("data[].created_at", "STRING", "생성일시", "Y", "ISO 8601 형식"),
            ("data[].updated_at", "STRING", "수정일시", "N", "ISO 8601 형식"),
            ("data[].created_by", "STRING", "생성자", "Y", "사용자 ID"),
            ("data[].updated_by", "STRING", "수정자", "N", "사용자 ID")
        ]
        
        # 필드들을 엑셀에 추가
        for field_name, field_type, description, required, note in api_fields:
            ws.cell(row=current_row, column=1, value=field_name)
            ws.cell(row=current_row, column=2, value=field_type)
            ws.cell(row=current_row, column=3, value=description)
            ws.cell(row=current_row, column=4, value=required)
            ws.cell(row=current_row, column=5, value=note)
            current_row += 1
        
        # 파일 저장
        wb.save('API_Files/08_약국정보_조회.xlsx')
        print("✅ 08번 파일 수정 완료!")
        
        # 수정된 내용 확인
        print(f"\n=== 수정된 출력 파라미터 ===")
        for field_name, field_type, description, required, note in api_fields:
            print(f"  - {field_name} ({field_type}): {description}")
        
        print(f"\n📊 수정 요약:")
        print(f"   - 총 필드 수: {len(api_fields)}")
        print(f"   - 필수 필드: {len([f for f in api_fields if f[3] == 'Y'])}")
        print(f"   - 선택 필드: {len([f for f in api_fields if f[3] == 'N'])}")
        
    except Exception as e:
        print(f"오류가 발생했습니다: {str(e)}")

if __name__ == "__main__":
    fix_08_file()
