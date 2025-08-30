#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import json

def create_api_excel():
    # API 데이터 정의
    api_data = [
        {
            "번호": 1,
            "URI": "/",
            "설명": "API 서버 상태 확인",
            "주기": "실시간",
            "메서드": "GET",
            "인증": "불필요",
            "입력파라미터": "없음",
            "출력파라미터": "name(STRING), version(STRING), status(STRING), timestamp(STRING), environment(STRING)"
        },
        {
            "번호": 2,
            "URI": "/health",
            "설명": "시스템 전반 상태 확인",
            "주기": "실시간",
            "메서드": "GET",
            "인증": "Bearer Token",
            "입력파라미터": "없음",
            "출력파라미터": "status(STRING), timestamp(STRING), uptime(FLOAT), environment(STRING)"
        },
        {
            "번호": 3,
            "URI": "/auth",
            "설명": "사용자 인증 및 토큰 발급",
            "주기": "실시간",
            "메서드": "POST",
            "인증": "불필요",
            "입력파라미터": "email(STRING), password(STRING)",
            "출력파라미터": "success(BOOLEAN), token(STRING), user(OBJECT), message(STRING)"
        },
        {
            "번호": 4,
            "URI": "/verify-token",
            "설명": "JWT 토큰 유효성 검증",
            "주기": "실시간",
            "메서드": "POST",
            "인증": "불필요",
            "입력파라미터": "token(STRING)",
            "출력파라미터": "success(BOOLEAN), valid(BOOLEAN), user(OBJECT), message(STRING)"
        },
        {
            "번호": 5,
            "URI": "/companies",
            "설명": "등록된 회사 정보 목록 조회",
            "주기": "실시간",
            "메서드": "GET",
            "인증": "Bearer Token",
            "입력파라미터": "page(INTEGER), limit(INTEGER), search(STRING), status(STRING)",
            "출력파라미터": "success(BOOLEAN), data(ARRAY), count(INTEGER), page(INTEGER), limit(INTEGER)"
        },
        {
            "번호": 6,
            "URI": "/products",
            "설명": "등록된 제품 정보 목록 조회",
            "주기": "실시간",
            "메서드": "GET",
            "인증": "Bearer Token",
            "입력파라미터": "page(INTEGER), limit(INTEGER), search(STRING), status(STRING)",
            "출력파라미터": "success(BOOLEAN), data(ARRAY), count(INTEGER), page(INTEGER), limit(INTEGER)"
        },
        {
            "번호": 7,
            "URI": "/clients",
            "설명": "등록된 병원 정보 목록 조회",
            "주기": "실시간",
            "메서드": "GET",
            "인증": "Bearer Token",
            "입력파라미터": "page(INTEGER), limit(INTEGER), search(STRING), status(STRING)",
            "출력파라미터": "success(BOOLEAN), data(ARRAY), count(INTEGER), page(INTEGER), limit(INTEGER)"
        },
        {
            "번호": 8,
            "URI": "/pharmacies",
            "설명": "등록된 약국 정보 목록 조회",
            "주기": "실시간",
            "메서드": "GET",
            "인증": "Bearer Token",
            "입력파라미터": "page(INTEGER), limit(INTEGER), search(STRING), status(STRING)",
            "출력파라미터": "success(BOOLEAN), data(ARRAY), count(INTEGER), page(INTEGER), limit(INTEGER)"
        },
        {
            "번호": 9,
            "URI": "/notices",
            "설명": "시스템 공지사항 목록 조회",
            "주기": "실시간",
            "메서드": "GET",
            "인증": "Bearer Token",
            "입력파라미터": "page(INTEGER), limit(INTEGER), search(STRING)",
            "출력파라미터": "success(BOOLEAN), data(ARRAY), count(INTEGER), page(INTEGER), limit(INTEGER)"
        },
        {
            "번호": 10,
            "URI": "/hospital-company-mappings",
            "설명": "병원과 업체 간의 관계 정보 조회",
            "주기": "실시간",
            "메서드": "GET",
            "인증": "Bearer Token",
            "입력파라미터": "page(INTEGER), limit(INTEGER), hospital_id(STRING), company_id(STRING)",
            "출력파라미터": "success(BOOLEAN), data(ARRAY), count(INTEGER), page(INTEGER), limit(INTEGER)"
        },
        {
            "번호": 11,
            "URI": "/hospital-pharmacy-mappings",
            "설명": "병원과 약국 간의 관계 정보 조회",
            "주기": "실시간",
            "메서드": "GET",
            "인증": "Bearer Token",
            "입력파라미터": "page(INTEGER), limit(INTEGER), hospital_id(STRING), pharmacy_id(STRING)",
            "출력파라미터": "success(BOOLEAN), data(ARRAY), count(INTEGER), page(INTEGER), limit(INTEGER)"
        },
        {
            "번호": 12,
            "URI": "/client-company-assignments",
            "설명": "병원과 업체 간의 배정 정보 조회",
            "주기": "실시간",
            "메서드": "GET",
            "인증": "Bearer Token",
            "입력파라미터": "page(INTEGER), limit(INTEGER), client_id(STRING), company_id(STRING)",
            "출력파라미터": "success(BOOLEAN), data(ARRAY), count(INTEGER), page(INTEGER), limit(INTEGER)"
        },
        {
            "번호": 13,
            "URI": "/client-pharmacy-assignments",
            "설명": "병원과 약국 간의 배정 정보 조회",
            "주기": "실시간",
            "메서드": "GET",
            "인증": "Bearer Token",
            "입력파라미터": "page(INTEGER), limit(INTEGER), client_id(STRING), pharmacy_id(STRING)",
            "출력파라미터": "success(BOOLEAN), data(ARRAY), count(INTEGER), page(INTEGER), limit(INTEGER)"
        },
        {
            "번호": 14,
            "URI": "/product-company-not-assignments",
            "설명": "업체에 배정되지 않은 제품 정보 조회",
            "주기": "실시간",
            "메서드": "GET",
            "인증": "Bearer Token",
            "입력파라미터": "page(INTEGER), limit(INTEGER), company_id(STRING)",
            "출력파라미터": "success(BOOLEAN), data(ARRAY), count(INTEGER), page(INTEGER), limit(INTEGER)"
        },
        {
            "번호": 15,
            "URI": "/wholesale-sales",
            "설명": "도매 매출 데이터 조회",
            "주기": "실시간",
            "메서드": "GET",
            "인증": "Bearer Token",
            "입력파라미터": "page(INTEGER), limit(INTEGER), start_date(STRING), end_date(STRING), pharmacy_code(STRING)",
            "출력파라미터": "success(BOOLEAN), data(ARRAY), count(INTEGER), page(INTEGER), limit(INTEGER)"
        },
        {
            "번호": 16,
            "URI": "/direct-sales",
            "설명": "직매 매출 데이터 조회",
            "주기": "실시간",
            "메서드": "GET",
            "인증": "Bearer Token",
            "입력파라미터": "page(INTEGER), limit(INTEGER), start_date(STRING), end_date(STRING), pharmacy_code(STRING)",
            "출력파라미터": "success(BOOLEAN), data(ARRAY), count(INTEGER), page(INTEGER), limit(INTEGER)"
        },
        {
            "번호": 17,
            "URI": "/performance-records",
            "설명": "실적 정보 목록 조회",
            "주기": "실시간",
            "메서드": "GET",
            "인증": "Bearer Token",
            "입력파라미터": "page(INTEGER), limit(INTEGER), settlement_month(STRING), company_id(STRING), client_id(STRING)",
            "출력파라미터": "success(BOOLEAN), data(ARRAY), count(INTEGER), page(INTEGER), limit(INTEGER)"
        },
        {
            "번호": 18,
            "URI": "/performance-records-absorption",
            "설명": "실적 흡수율 분석 정보 조회",
            "주기": "실시간",
            "메서드": "GET",
            "인증": "Bearer Token",
            "입력파라미터": "page(INTEGER), limit(INTEGER), settlement_month(STRING), company_id(STRING), client_id(STRING)",
            "출력파라미터": "success(BOOLEAN), data(ARRAY), count(INTEGER), page(INTEGER), limit(INTEGER)"
        },
        {
            "번호": 19,
            "URI": "/performance-evidence-files",
            "설명": "실적 증빙 파일 정보 조회",
            "주기": "실시간",
            "메서드": "GET",
            "인증": "Bearer Token",
            "입력파라미터": "page(INTEGER), limit(INTEGER), settlement_month(STRING), company_id(STRING), client_id(STRING)",
            "출력파라미터": "success(BOOLEAN), data(ARRAY), count(INTEGER), page(INTEGER), limit(INTEGER)"
        },
        {
            "번호": 20,
            "URI": "/settlement-months",
            "설명": "정산월 목록 조회",
            "주기": "실시간",
            "메서드": "GET",
            "인증": "Bearer Token",
            "입력파라미터": "page(INTEGER), limit(INTEGER), status(STRING)",
            "출력파라미터": "success(BOOLEAN), data(ARRAY), count(INTEGER), page(INTEGER), limit(INTEGER)"
        },
        {
            "번호": 21,
            "URI": "/settlement-share",
            "설명": "정산내역서 목록 조회",
            "주기": "실시간",
            "메서드": "GET",
            "인증": "Bearer Token",
            "입력파라미터": "page(INTEGER), limit(INTEGER), settlement_month(STRING), company_id(STRING)",
            "출력파라미터": "success(BOOLEAN), data(ARRAY), count(INTEGER), page(INTEGER), limit(INTEGER)"
        }
    ]

    # DataFrame 생성
    df = pd.DataFrame(api_data)
    
    # Excel 파일 생성
    filename = "Sinil_PMS_API_Endpoints.xlsx"
    
    # ExcelWriter 사용하여 스타일링 적용
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='API_Endpoints', index=False)
        
        # 워크시트 가져오기
        worksheet = writer.sheets['API_Endpoints']
        
        # 스타일 정의
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        wrap_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # 헤더 스타일 적용
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        # 데이터 행 스타일 적용
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.border = border
                if cell.column in [1, 2, 5, 6]:  # 번호, URI, 메서드, 인증 컬럼
                    cell.alignment = center_alignment
                else:  # 설명, 주기, 입력파라미터, 출력파라미터 컬럼
                    cell.alignment = wrap_alignment
        
        # 컬럼 너비 조정
        column_widths = {
            'A': 8,   # 번호
            'B': 35,  # URI
            'C': 40,  # 설명
            'D': 12,  # 주기
            'E': 10,  # 메서드
            'F': 15,  # 인증
            'G': 50,  # 입력파라미터
            'H': 60   # 출력파라미터
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # 행 높이 조정
        for row in range(2, worksheet.max_row + 1):
            worksheet.row_dimensions[row].height = 60
        
        # 헤더 행 높이 조정
        worksheet.row_dimensions[1].height = 30
    
    print(f"✅ Excel 파일이 생성되었습니다: {filename}")
    print(f"📊 총 {len(api_data)}개의 API 엔드포인트가 포함되었습니다.")
    
    return filename

if __name__ == "__main__":
    create_api_excel()
