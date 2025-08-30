#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
from openpyxl import load_workbook
import requests
import json

def check_09_file():
    """09번 파일을 확인합니다."""
    
    try:
        print("=== 09번 파일 (공지사항_조회) 확인 ===")
        
        # 먼저 로그인해서 토큰을 받아옵니다
        print("로그인하여 토큰 받는 중...")
        login_data = {
            "email": "admin@admin.com",
            "password": "asdf1234"
        }
        
        login_response = requests.post('https://shinil.vercel.app/api/auth', json=login_data)
        
        if login_response.status_code != 200:
            print(f"로그인 실패: {login_response.status_code}")
            return
        
        login_result = login_response.json()
        token = login_result.get('data', {}).get('token')
        
        if not token:
            print("토큰을 받을 수 없습니다.")
            return
        
        print("토큰 받기 성공!")
        
        # 실제 API 응답 확인 (공지사항 조회)
        print("실제 API 응답 확인 중...")
        headers = {
            'Authorization': f'Bearer {token}'
        }
        
        response = requests.get('https://shinil.vercel.app/api/notices', headers=headers)
        
        if response.status_code == 200:
            actual_response = response.json()
            print("실제 API 응답:")
            print(json.dumps(actual_response, indent=2, ensure_ascii=False))
            
            # 실제 필드들 추출 (중첩 구조 포함)
            actual_fields = []
            for key, value in actual_response.items():
                if key == "data" and isinstance(value, list) and len(value) > 0:
                    # data 배열의 첫 번째 항목의 필드들을 추출
                    first_item = value[0]
                    for sub_key in first_item.keys():
                        actual_fields.append(f"data[].{sub_key}")
                else:
                    actual_fields.append(key)
            
            print(f"\n실제 API 응답 필드들: {actual_fields}")
        else:
            print(f"API 요청 실패: {response.status_code}")
            print(f"응답 내용: {response.text}")
            return
        
        # 엑셀 파일 확인
        print(f"\n=== 09번 엑셀 파일 내용 ===")
        wb = load_workbook('API_Files/09_공지사항_조회.xlsx')
        ws = wb.active
        
        # 출력 파라미터 섹션 찾기
        output_section_row = None
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and "출력 파라미터" in str(cell_value):
                output_section_row = row
                break
        
        if output_section_row:
            # 출력 파라미터 섹션의 끝 찾기
            output_end_row = output_section_row
            for row in range(output_section_row + 1, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=1).value
                if cell_value and any(keyword in str(cell_value) for keyword in ["입력 파라미터", "응답 예시"]):
                    break
                output_end_row = row
            
            print(f"출력 파라미터 섹션: 행 {output_section_row} ~ {output_end_row}")
            
            # 출력 파라미터 내용 출력 (처음 15행만)
            for row in range(output_section_row, min(output_end_row + 1, output_section_row + 15)):
                row_data = []
                for col in range(1, 6):
                    cell_value = ws.cell(row=row, column=col).value
                    row_data.append(str(cell_value) if cell_value is not None else "")
                
                if any(row_data):
                    print(f"  행 {row}: {' | '.join(row_data)}")
            
            # 엑셀 필드명들 추출
            excel_fields = []
            for row in range(output_section_row + 2, output_end_row + 1):
                field_name = ws.cell(row=row, column=1).value
                if field_name and field_name != "출력 파라미터":
                    excel_fields.append(field_name)
            
            print(f"\n엑셀 파일의 출력 파라미터 필드들: {excel_fields}")
        
        # 비교
        print(f"\n=== 비교 결과 ===")
        
        # API 응답에는 있지만 엑셀에 없는 필드들
        missing_in_excel = set(actual_fields) - set(excel_fields)
        if missing_in_excel:
            print(f"❌ API 응답에는 있지만 엑셀에 누락된 필드들:")
            for field in missing_in_excel:
                print(f"   - {field}")
        
        # 엑셀에는 있지만 API 응답에 없는 필드들
        missing_in_api = set(excel_fields) - set(actual_fields)
        if missing_in_api:
            print(f"❌ 엑셀에는 있지만 API 응답에 없는 필드들:")
            for field in missing_in_api:
                print(f"   - {field}")
        
        # 일치하는 필드들
        matching_fields = set(actual_fields) & set(excel_fields)
        if matching_fields:
            print(f"✅ 일치하는 필드들:")
            for field in matching_fields:
                print(f"   - {field}")
        
        if not missing_in_excel and not missing_in_api:
            print("✅ 모든 필드가 일치합니다!")
        else:
            print(f"\n📊 요약:")
            print(f"   - 실제 API 필드 수: {len(actual_fields)}")
            print(f"   - 엑셀 파일 필드 수: {len(excel_fields)}")
            print(f"   - 일치하는 필드 수: {len(matching_fields)}")
            print(f"   - 누락된 필드 수: {len(missing_in_excel)}")
            print(f"   - 불필요한 필드 수: {len(missing_in_api)}")
        
    except Exception as e:
        print(f"오류가 발생했습니다: {str(e)}")

if __name__ == "__main__":
    check_09_file()
