#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

def create_api_file(api_data, filename):
    """개별 API 파일 생성"""
    wb = Workbook()
    ws = wb.active
    
    # 스타일 정의
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    wrap_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    current_row = 1
    
    # API 기본 정보 테이블
    info_headers = ["항목", "내용"]
    for col, header in enumerate(info_headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    current_row += 1
    
    # API 정보 데이터
    api_info = [
        ["URI", api_data["uri"]],
        ["설명", api_data["description"]],
        ["주기", api_data["frequency"]],
        ["메서드", api_data["method"]],
        ["인증", api_data["auth"]]
    ]
    
    for info in api_info:
        for col, value in enumerate(info, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.border = border
            if col == 1:
                cell.font = Font(bold=True)
                cell.alignment = center_alignment
            else:
                cell.alignment = wrap_alignment
        current_row += 1
    
    current_row += 2
    
    # 입력 파라미터 테이블
    if api_data["input_params"]:
        ws.cell(row=current_row, column=1, value="입력 파라미터").font = Font(bold=True, size=12)
        current_row += 1
        
        input_headers = ["파라미터명", "타입", "설명", "필수"]
        for col, header in enumerate(input_headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        current_row += 1
        
        for param in api_data["input_params"]:
            for col, value in enumerate([param["name"], param["type"], param["description"], param["required"]], 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.border = border
                cell.alignment = wrap_alignment
            current_row += 1
    else:
        ws.cell(row=current_row, column=1, value="입력 파라미터").font = Font(bold=True, size=12)
        current_row += 1
        ws.cell(row=current_row, column=1, value="없음").border = border
        current_row += 1
    
    current_row += 2
    
    # 출력 파라미터 테이블
    ws.cell(row=current_row, column=1, value="출력 파라미터").font = Font(bold=True, size=12)
    current_row += 1
    
    output_headers = ["파라미터명", "타입", "설명", "필수", "비고"]
    for col, header in enumerate(output_headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    current_row += 1
    
    for param in api_data["output_params"]:
        for col, value in enumerate([param["name"], param["type"], param["description"], param["required"], param["remarks"]], 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.border = border
            cell.alignment = wrap_alignment
        current_row += 1
    
    # 컬럼 너비 조정
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15
    
    # 파일 저장
    wb.save(filename)
    print(f"✅ 생성됨: {filename}")

def create_all_api_files():
    """모든 API 파일 생성"""
    
    # API 데이터 정의 (전체 21개)
    apis = [
        {
            "name": "01_API_상태확인",
            "uri": "/",
            "description": "API 서버 상태 확인",
            "frequency": "실시간",
            "method": "GET",
            "auth": "불필요",
            "input_params": [],
            "output_params": [
                {"name": "name", "type": "STRING", "description": "API 서버명", "required": "Y", "remarks": ""},
                {"name": "version", "type": "STRING", "description": "API 버전", "required": "Y", "remarks": ""},
                {"name": "status", "type": "STRING", "description": "서버 상태", "required": "Y", "remarks": ""},
                {"name": "timestamp", "type": "STRING", "description": "응답 시간", "required": "Y", "remarks": ""},
                {"name": "environment", "type": "STRING", "description": "환경 정보", "required": "Y", "remarks": ""}
            ]
        },
        {
            "name": "02_시스템_헬스체크",
            "uri": "/health",
            "description": "시스템 전반 상태 확인",
            "frequency": "실시간",
            "method": "GET",
            "auth": "Bearer Token",
            "input_params": [],
            "output_params": [
                {"name": "status", "type": "STRING", "description": "시스템 상태", "required": "Y", "remarks": ""},
                {"name": "timestamp", "type": "STRING", "description": "응답 시간", "required": "Y", "remarks": ""},
                {"name": "uptime", "type": "FLOAT", "description": "서버 가동시간(초)", "required": "Y", "remarks": ""},
                {"name": "environment", "type": "STRING", "description": "환경 정보", "required": "Y", "remarks": ""}
            ]
        },
        {
            "name": "03_사용자_로그인",
            "uri": "/auth",
            "description": "사용자 인증 및 토큰 발급",
            "frequency": "실시간",
            "method": "POST",
            "auth": "불필요",
            "input_params": [
                {"name": "email", "type": "STRING", "description": "사용자 이메일", "required": "Y", "remarks": ""},
                {"name": "password", "type": "STRING", "description": "사용자 비밀번호", "required": "Y", "remarks": ""}
            ],
            "output_params": [
                {"name": "success", "type": "BOOLEAN", "description": "성공 여부", "required": "Y", "remarks": ""},
                {"name": "token", "type": "STRING", "description": "JWT 토큰", "required": "Y", "remarks": ""},
                {"name": "user", "type": "OBJECT", "description": "사용자 정보", "required": "Y", "remarks": ""},
                {"name": "message", "type": "STRING", "description": "응답 메시지", "required": "Y", "remarks": ""}
            ]
        },
        {
            "name": "04_토큰_검증",
            "uri": "/verify-token",
            "description": "JWT 토큰 유효성 검증",
            "frequency": "실시간",
            "method": "POST",
            "auth": "불필요",
            "input_params": [
                {"name": "token", "type": "STRING", "description": "검증할 JWT 토큰", "required": "Y", "remarks": ""}
            ],
            "output_params": [
                {"name": "success", "type": "BOOLEAN", "description": "성공 여부", "required": "Y", "remarks": ""},
                {"name": "valid", "type": "BOOLEAN", "description": "토큰 유효성", "required": "Y", "remarks": ""},
                {"name": "user", "type": "OBJECT", "description": "사용자 정보", "required": "N", "remarks": "토큰이 유효한 경우"},
                {"name": "message", "type": "STRING", "description": "응답 메시지", "required": "Y", "remarks": ""}
            ]
        },
        {
            "name": "05_회사정보_조회",
            "uri": "/companies",
            "description": "등록된 회사 정보 목록 조회",
            "frequency": "실시간",
            "method": "GET",
            "auth": "Bearer Token",
            "input_params": [
                {"name": "page", "type": "INTEGER", "description": "페이지 번호", "required": "N", "remarks": ""},
                {"name": "limit", "type": "INTEGER", "description": "페이지당 항목 수", "required": "N", "remarks": ""},
                {"name": "search", "type": "STRING", "description": "검색어", "required": "N", "remarks": ""},
                {"name": "status", "type": "STRING", "description": "상태 필터", "required": "N", "remarks": ""}
            ],
            "output_params": [
                {"name": "success", "type": "BOOLEAN", "description": "성공 여부", "required": "Y", "remarks": ""},
                {"name": "data", "type": "ARRAY", "description": "회사 정보 목록", "required": "Y", "remarks": ""},
                {"name": "data[].id", "type": "UUID", "description": "회사 ID", "required": "Y", "remarks": ""},
                {"name": "data[].user_id", "type": "UUID", "description": "연결된 사용자 ID", "required": "N", "remarks": ""},
                {"name": "data[].company_name", "type": "STRING", "description": "회사명", "required": "Y", "remarks": ""},
                {"name": "data[].business_registration_number", "type": "STRING", "description": "사업자등록번호", "required": "Y", "remarks": ""},
                {"name": "data[].representative_name", "type": "STRING", "description": "대표자명", "required": "Y", "remarks": ""},
                {"name": "data[].business_address", "type": "STRING", "description": "사업장 주소", "required": "Y", "remarks": ""},
                {"name": "data[].landline_phone", "type": "STRING", "description": "대표전화", "required": "N", "remarks": ""},
                {"name": "data[].contact_person_name", "type": "STRING", "description": "담당자명", "required": "Y", "remarks": ""},
                {"name": "data[].mobile_phone", "type": "STRING", "description": "휴대폰 번호", "required": "Y", "remarks": ""},
                {"name": "data[].mobile_phone_2", "type": "STRING", "description": "휴대폰 번호 2", "required": "N", "remarks": ""},
                {"name": "data[].email", "type": "STRING", "description": "이메일", "required": "Y", "remarks": ""},
                {"name": "data[].default_commission_grade", "type": "STRING", "description": "기본 수수료 등급", "required": "Y", "remarks": "A/B/C"},
                {"name": "data[].remarks", "type": "STRING", "description": "비고", "required": "N", "remarks": ""},
                {"name": "data[].approval_status", "type": "STRING", "description": "승인 상태", "required": "Y", "remarks": "pending/approved/rejected"},
                {"name": "data[].status", "type": "STRING", "description": "상태", "required": "Y", "remarks": "active/inactive"},
                {"name": "data[].user_type", "type": "STRING", "description": "사용자 타입", "required": "Y", "remarks": "user/admin"},
                {"name": "data[].company_group", "type": "STRING", "description": "회사 그룹", "required": "N", "remarks": ""},
                {"name": "data[].assigned_pharmacist_contact", "type": "STRING", "description": "배정 약사 연락처", "required": "N", "remarks": ""},
                {"name": "data[].receive_email", "type": "STRING", "description": "수신 이메일", "required": "N", "remarks": ""},
                {"name": "data[].created_at", "type": "TIMESTAMP", "description": "생성일시", "required": "Y", "remarks": ""},
                {"name": "data[].updated_at", "type": "TIMESTAMP", "description": "수정일시", "required": "Y", "remarks": ""},
                {"name": "data[].approved_at", "type": "TIMESTAMP", "description": "승인일시", "required": "N", "remarks": ""},
                {"name": "count", "type": "INTEGER", "description": "전체 항목 수", "required": "Y", "remarks": ""},
                {"name": "page", "type": "INTEGER", "description": "현재 페이지", "required": "Y", "remarks": ""},
                {"name": "limit", "type": "INTEGER", "description": "페이지당 항목 수", "required": "Y", "remarks": ""}
            ]
        },
        {
            "name": "06_제품정보_조회",
            "uri": "/products",
            "description": "등록된 제품 정보 목록 조회",
            "frequency": "실시간",
            "method": "GET",
            "auth": "Bearer Token",
            "input_params": [
                {"name": "page", "type": "INTEGER", "description": "페이지 번호", "required": "N", "remarks": ""},
                {"name": "limit", "type": "INTEGER", "description": "페이지당 항목 수", "required": "N", "remarks": ""},
                {"name": "search", "type": "STRING", "description": "검색어", "required": "N", "remarks": ""},
                {"name": "status", "type": "STRING", "description": "상태 필터", "required": "N", "remarks": ""}
            ],
            "output_params": [
                {"name": "success", "type": "BOOLEAN", "description": "성공 여부", "required": "Y", "remarks": ""},
                {"name": "data", "type": "ARRAY", "description": "제품 정보 목록", "required": "Y", "remarks": ""},
                {"name": "data[].id", "type": "UUID", "description": "제품 ID", "required": "Y", "remarks": ""},
                {"name": "data[].product_name", "type": "STRING", "description": "제품명", "required": "Y", "remarks": ""},
                {"name": "data[].insurance_code", "type": "STRING", "description": "보험 코드", "required": "N", "remarks": ""},
                {"name": "data[].price", "type": "INTEGER", "description": "가격", "required": "N", "remarks": ""},
                {"name": "data[].commission_rate_a", "type": "NUMERIC", "description": "A등급 수수료율", "required": "Y", "remarks": ""},
                {"name": "data[].commission_rate_b", "type": "NUMERIC", "description": "B등급 수수료율", "required": "Y", "remarks": ""},
                {"name": "data[].commission_rate_c", "type": "NUMERIC", "description": "C등급 수수료율", "required": "N", "remarks": ""},
                {"name": "data[].standard_code", "type": "STRING", "description": "표준 코드", "required": "Y", "remarks": ""},
                {"name": "data[].unit_packaging_desc", "type": "STRING", "description": "단위 포장 설명", "required": "N", "remarks": ""},
                {"name": "data[].unit_quantity", "type": "INTEGER", "description": "단위 수량", "required": "N", "remarks": ""},
                {"name": "data[].remarks", "type": "STRING", "description": "비고", "required": "N", "remarks": ""},
                {"name": "data[].status", "type": "STRING", "description": "상태", "required": "Y", "remarks": "active/inactive"},
                {"name": "data[].base_month", "type": "STRING", "description": "기준월", "required": "Y", "remarks": ""},
                {"name": "data[].created_at", "type": "TIMESTAMP", "description": "생성일시", "required": "Y", "remarks": ""},
                {"name": "data[].updated_at", "type": "TIMESTAMP", "description": "수정일시", "required": "Y", "remarks": ""},
                {"name": "count", "type": "INTEGER", "description": "전체 항목 수", "required": "Y", "remarks": ""},
                {"name": "page", "type": "INTEGER", "description": "현재 페이지", "required": "Y", "remarks": ""},
                {"name": "limit", "type": "INTEGER", "description": "페이지당 항목 수", "required": "Y", "remarks": ""}
            ]
        },
        {
            "name": "07_병원정보_조회",
            "uri": "/clients",
            "description": "등록된 병원 정보 목록 조회",
            "frequency": "실시간",
            "method": "GET",
            "auth": "Bearer Token",
            "input_params": [
                {"name": "page", "type": "INTEGER", "description": "페이지 번호", "required": "N", "remarks": ""},
                {"name": "limit", "type": "INTEGER", "description": "페이지당 항목 수", "required": "N", "remarks": ""},
                {"name": "search", "type": "STRING", "description": "검색어", "required": "N", "remarks": ""},
                {"name": "status", "type": "STRING", "description": "상태 필터", "required": "N", "remarks": ""}
            ],
            "output_params": [
                {"name": "success", "type": "BOOLEAN", "description": "성공 여부", "required": "Y", "remarks": ""},
                {"name": "data", "type": "ARRAY", "description": "병원 정보 목록", "required": "Y", "remarks": ""},
                {"name": "data[].id", "type": "BIGINT", "description": "병원 ID", "required": "Y", "remarks": ""},
                {"name": "data[].client_code", "type": "STRING", "description": "고객 코드", "required": "N", "remarks": ""},
                {"name": "data[].name", "type": "STRING", "description": "병원명", "required": "Y", "remarks": ""},
                {"name": "data[].business_registration_number", "type": "STRING", "description": "사업자등록번호", "required": "Y", "remarks": ""},
                {"name": "data[].owner_name", "type": "STRING", "description": "대표자명", "required": "N", "remarks": ""},
                {"name": "data[].address", "type": "STRING", "description": "주소", "required": "N", "remarks": ""},
                {"name": "data[].remarks", "type": "STRING", "description": "비고", "required": "N", "remarks": ""},
                {"name": "data[].status", "type": "STRING", "description": "상태", "required": "Y", "remarks": "active/inactive"},
                {"name": "data[].created_at", "type": "TIMESTAMP", "description": "생성일시", "required": "Y", "remarks": ""},
                {"name": "data[].updated_at", "type": "TIMESTAMP", "description": "수정일시", "required": "Y", "remarks": ""},
                {"name": "count", "type": "INTEGER", "description": "전체 항목 수", "required": "Y", "remarks": ""},
                {"name": "page", "type": "INTEGER", "description": "현재 페이지", "required": "Y", "remarks": ""},
                {"name": "limit", "type": "INTEGER", "description": "페이지당 항목 수", "required": "Y", "remarks": ""}
            ]
        },
        {
            "name": "08_약국정보_조회",
            "uri": "/pharmacies",
            "description": "등록된 약국 정보 목록 조회",
            "frequency": "실시간",
            "method": "GET",
            "auth": "Bearer Token",
            "input_params": [
                {"name": "page", "type": "INTEGER", "description": "페이지 번호", "required": "N", "remarks": ""},
                {"name": "limit", "type": "INTEGER", "description": "페이지당 항목 수", "required": "N", "remarks": ""},
                {"name": "search", "type": "STRING", "description": "검색어", "required": "N", "remarks": ""},
                {"name": "status", "type": "STRING", "description": "상태 필터", "required": "N", "remarks": ""}
            ],
            "output_params": [
                {"name": "success", "type": "BOOLEAN", "description": "성공 여부", "required": "Y", "remarks": ""},
                {"name": "data", "type": "ARRAY", "description": "약국 정보 목록", "required": "Y", "remarks": ""},
                {"name": "data[].id", "type": "BIGINT", "description": "약국 ID", "required": "Y", "remarks": ""},
                {"name": "data[].pharmacy_code", "type": "STRING", "description": "약국 코드", "required": "N", "remarks": ""},
                {"name": "data[].name", "type": "STRING", "description": "약국명", "required": "Y", "remarks": ""},
                {"name": "data[].business_registration_number", "type": "STRING", "description": "사업자등록번호", "required": "Y", "remarks": ""},
                {"name": "data[].address", "type": "STRING", "description": "주소", "required": "N", "remarks": ""},
                {"name": "data[].remarks", "type": "STRING", "description": "비고", "required": "N", "remarks": ""},
                {"name": "data[].status", "type": "STRING", "description": "상태", "required": "Y", "remarks": "active/inactive"},
                {"name": "data[].created_at", "type": "TIMESTAMP", "description": "생성일시", "required": "Y", "remarks": ""},
                {"name": "data[].updated_at", "type": "TIMESTAMP", "description": "수정일시", "required": "Y", "remarks": ""},
                {"name": "count", "type": "INTEGER", "description": "전체 항목 수", "required": "Y", "remarks": ""},
                {"name": "page", "type": "INTEGER", "description": "현재 페이지", "required": "Y", "remarks": ""},
                {"name": "limit", "type": "INTEGER", "description": "페이지당 항목 수", "required": "Y", "remarks": ""}
            ]
        },
        {
            "name": "09_공지사항_조회",
            "uri": "/notices",
            "description": "시스템 공지사항 목록 조회",
            "frequency": "실시간",
            "method": "GET",
            "auth": "Bearer Token",
            "input_params": [
                {"name": "page", "type": "INTEGER", "description": "페이지 번호", "required": "N", "remarks": ""},
                {"name": "limit", "type": "INTEGER", "description": "페이지당 항목 수", "required": "N", "remarks": ""},
                {"name": "search", "type": "STRING", "description": "검색어", "required": "N", "remarks": ""}
            ],
            "output_params": [
                {"name": "success", "type": "BOOLEAN", "description": "성공 여부", "required": "Y", "remarks": ""},
                {"name": "data", "type": "ARRAY", "description": "공지사항 목록", "required": "Y", "remarks": ""},
                {"name": "data[].id", "type": "UUID", "description": "공지사항 ID", "required": "Y", "remarks": ""},
                {"name": "data[].title", "type": "STRING", "description": "제목", "required": "Y", "remarks": ""},
                {"name": "data[].content", "type": "STRING", "description": "내용", "required": "Y", "remarks": ""},
                {"name": "data[].is_pinned", "type": "BOOLEAN", "description": "고정 여부", "required": "Y", "remarks": ""},
                {"name": "data[].view_count", "type": "INTEGER", "description": "조회수", "required": "Y", "remarks": ""},
                {"name": "data[].author_id", "type": "UUID", "description": "작성자 ID", "required": "N", "remarks": ""},
                {"name": "data[].file_url", "type": "STRING", "description": "첨부파일 URL", "required": "N", "remarks": ""},
                {"name": "data[].links", "type": "STRING", "description": "링크", "required": "N", "remarks": ""},
                {"name": "data[].created_at", "type": "TIMESTAMP", "description": "생성일시", "required": "Y", "remarks": ""},
                {"name": "data[].updated_at", "type": "TIMESTAMP", "description": "수정일시", "required": "Y", "remarks": ""},
                {"name": "count", "type": "INTEGER", "description": "전체 항목 수", "required": "Y", "remarks": ""},
                {"name": "page", "type": "INTEGER", "description": "현재 페이지", "required": "Y", "remarks": ""},
                {"name": "limit", "type": "INTEGER", "description": "페이지당 항목 수", "required": "Y", "remarks": ""}
            ]
        },
        {
            "name": "10_병원업체_관계정보",
            "uri": "/hospital-company-mappings",
            "description": "병원과 업체 간의 관계 정보 조회",
            "frequency": "실시간",
            "method": "GET",
            "auth": "Bearer Token",
            "input_params": [
                {"name": "page", "type": "INTEGER", "description": "페이지 번호", "required": "N", "remarks": ""},
                {"name": "limit", "type": "INTEGER", "description": "페이지당 항목 수", "required": "N", "remarks": ""},
                {"name": "hospital_id", "type": "STRING", "description": "병원 ID", "required": "N", "remarks": ""},
                {"name": "company_id", "type": "STRING", "description": "업체 ID", "required": "N", "remarks": ""}
            ],
            "output_params": [
                {"name": "success", "type": "BOOLEAN", "description": "성공 여부", "required": "Y", "remarks": ""},
                {"name": "data", "type": "ARRAY", "description": "관계 정보 목록", "required": "Y", "remarks": ""},
                {"name": "count", "type": "INTEGER", "description": "전체 항목 수", "required": "Y", "remarks": ""},
                {"name": "page", "type": "INTEGER", "description": "현재 페이지", "required": "Y", "remarks": ""},
                {"name": "limit", "type": "INTEGER", "description": "페이지당 항목 수", "required": "Y", "remarks": ""}
            ]
        },
        {
            "name": "11_병원약국_관계정보",
            "uri": "/hospital-pharmacy-mappings",
            "description": "병원과 약국 간의 관계 정보 조회",
            "frequency": "실시간",
            "method": "GET",
            "auth": "Bearer Token",
            "input_params": [
                {"name": "page", "type": "INTEGER", "description": "페이지 번호", "required": "N", "remarks": ""},
                {"name": "limit", "type": "INTEGER", "description": "페이지당 항목 수", "required": "N", "remarks": ""},
                {"name": "hospital_id", "type": "STRING", "description": "병원 ID", "required": "N", "remarks": ""},
                {"name": "pharmacy_id", "type": "STRING", "description": "약국 ID", "required": "N", "remarks": ""}
            ],
            "output_params": [
                {"name": "success", "type": "BOOLEAN", "description": "성공 여부", "required": "Y", "remarks": ""},
                {"name": "data", "type": "ARRAY", "description": "관계 정보 목록", "required": "Y", "remarks": ""},
                {"name": "count", "type": "INTEGER", "description": "전체 항목 수", "required": "Y", "remarks": ""},
                {"name": "page", "type": "INTEGER", "description": "현재 페이지", "required": "Y", "remarks": ""},
                {"name": "limit", "type": "INTEGER", "description": "페이지당 항목 수", "required": "Y", "remarks": ""}
            ]
        },
        {
            "name": "12_병원업체_매핑정보",
            "uri": "/client-company-assignments",
            "description": "병원과 업체 간의 배정 정보 조회",
            "frequency": "실시간",
            "method": "GET",
            "auth": "Bearer Token",
            "input_params": [
                {"name": "page", "type": "INTEGER", "description": "페이지 번호", "required": "N", "remarks": ""},
                {"name": "limit", "type": "INTEGER", "description": "페이지당 항목 수", "required": "N", "remarks": ""},
                {"name": "client_id", "type": "STRING", "description": "고객 ID", "required": "N", "remarks": ""},
                {"name": "company_id", "type": "STRING", "description": "업체 ID", "required": "N", "remarks": ""}
            ],
            "output_params": [
                {"name": "success", "type": "BOOLEAN", "description": "성공 여부", "required": "Y", "remarks": ""},
                {"name": "data", "type": "ARRAY", "description": "배정 정보 목록", "required": "Y", "remarks": ""},
                {"name": "count", "type": "INTEGER", "description": "전체 항목 수", "required": "Y", "remarks": ""},
                {"name": "page", "type": "INTEGER", "description": "현재 페이지", "required": "Y", "remarks": ""},
                {"name": "limit", "type": "INTEGER", "description": "페이지당 항목 수", "required": "Y", "remarks": ""}
            ]
        },
        {
            "name": "13_병원약국_매핑정보",
            "uri": "/client-pharmacy-assignments",
            "description": "병원과 약국 간의 배정 정보 조회",
            "frequency": "실시간",
            "method": "GET",
            "auth": "Bearer Token",
            "input_params": [
                {"name": "page", "type": "INTEGER", "description": "페이지 번호", "required": "N", "remarks": ""},
                {"name": "limit", "type": "INTEGER", "description": "페이지당 항목 수", "required": "N", "remarks": ""},
                {"name": "client_id", "type": "STRING", "description": "고객 ID", "required": "N", "remarks": ""},
                {"name": "pharmacy_id", "type": "STRING", "description": "약국 ID", "required": "N", "remarks": ""}
            ],
            "output_params": [
                {"name": "success", "type": "BOOLEAN", "description": "성공 여부", "required": "Y", "remarks": ""},
                {"name": "data", "type": "ARRAY", "description": "배정 정보 목록", "required": "Y", "remarks": ""},
                {"name": "count", "type": "INTEGER", "description": "전체 항목 수", "required": "Y", "remarks": ""},
                {"name": "page", "type": "INTEGER", "description": "현재 페이지", "required": "Y", "remarks": ""},
                {"name": "limit", "type": "INTEGER", "description": "페이지당 항목 수", "required": "Y", "remarks": ""}
            ]
        },
        {
            "name": "14_제품업체_미배정매핑",
            "uri": "/product-company-not-assignments",
            "description": "업체에 배정되지 않은 제품 정보 조회",
            "frequency": "실시간",
            "method": "GET",
            "auth": "Bearer Token",
            "input_params": [
                {"name": "page", "type": "INTEGER", "description": "페이지 번호", "required": "N", "remarks": ""},
                {"name": "limit", "type": "INTEGER", "description": "페이지당 항목 수", "required": "N", "remarks": ""},
                {"name": "company_id", "type": "STRING", "description": "업체 ID", "required": "N", "remarks": ""}
            ],
            "output_params": [
                {"name": "success", "type": "BOOLEAN", "description": "성공 여부", "required": "Y", "remarks": ""},
                {"name": "data", "type": "ARRAY", "description": "미배정 제품 목록", "required": "Y", "remarks": ""},
                {"name": "count", "type": "INTEGER", "description": "전체 항목 수", "required": "Y", "remarks": ""},
                {"name": "page", "type": "INTEGER", "description": "현재 페이지", "required": "Y", "remarks": ""},
                {"name": "limit", "type": "INTEGER", "description": "페이지당 항목 수", "required": "Y", "remarks": ""}
            ]
        },
        {
            "name": "15_도매매출_조회",
            "uri": "/wholesale-sales",
            "description": "도매 매출 데이터 조회",
            "frequency": "실시간",
            "method": "GET",
            "auth": "Bearer Token",
            "input_params": [
                {"name": "page", "type": "INTEGER", "description": "페이지 번호", "required": "N", "remarks": ""},
                {"name": "limit", "type": "INTEGER", "description": "페이지당 항목 수", "required": "N", "remarks": ""},
                {"name": "start_date", "type": "STRING", "description": "조회 시작일", "required": "N", "remarks": ""},
                {"name": "end_date", "type": "STRING", "description": "조회 종료일", "required": "N", "remarks": ""},
                {"name": "pharmacy_code", "type": "STRING", "description": "약국 코드", "required": "N", "remarks": ""}
            ],
            "output_params": [
                {"name": "success", "type": "BOOLEAN", "description": "성공 여부", "required": "Y", "remarks": ""},
                {"name": "data", "type": "ARRAY", "description": "도매 매출 목록", "required": "Y", "remarks": ""},
                {"name": "data[].id", "type": "BIGINT", "description": "매출 ID", "required": "Y", "remarks": ""},
                {"name": "data[].pharmacy_code", "type": "STRING", "description": "약국 코드", "required": "N", "remarks": ""},
                {"name": "data[].pharmacy_name", "type": "STRING", "description": "약국명", "required": "N", "remarks": ""},
                {"name": "data[].business_registration_number", "type": "STRING", "description": "사업자등록번호", "required": "Y", "remarks": ""},
                {"name": "data[].address", "type": "STRING", "description": "주소", "required": "N", "remarks": ""},
                {"name": "data[].standard_code", "type": "STRING", "description": "표준 코드", "required": "Y", "remarks": ""},
                {"name": "data[].product_name", "type": "STRING", "description": "제품명", "required": "N", "remarks": ""},
                {"name": "data[].sales_amount", "type": "NUMERIC", "description": "매출액", "required": "N", "remarks": ""},
                {"name": "data[].sales_date", "type": "DATE", "description": "매출일", "required": "N", "remarks": ""},
                {"name": "data[].created_by", "type": "STRING", "description": "생성자", "required": "N", "remarks": ""},
                {"name": "data[].updated_by", "type": "STRING", "description": "수정자", "required": "N", "remarks": ""},
                {"name": "data[].created_at", "type": "TIMESTAMP", "description": "생성일시", "required": "Y", "remarks": ""},
                {"name": "data[].updated_at", "type": "TIMESTAMP", "description": "수정일시", "required": "Y", "remarks": ""},
                {"name": "count", "type": "INTEGER", "description": "전체 항목 수", "required": "Y", "remarks": ""},
                {"name": "page", "type": "INTEGER", "description": "현재 페이지", "required": "Y", "remarks": ""},
                {"name": "limit", "type": "INTEGER", "description": "페이지당 항목 수", "required": "Y", "remarks": ""}
            ]
        },
        {
            "name": "16_직매매출_조회",
            "uri": "/direct-sales",
            "description": "직매 매출 데이터 조회",
            "frequency": "실시간",
            "method": "GET",
            "auth": "Bearer Token",
            "input_params": [
                {"name": "page", "type": "INTEGER", "description": "페이지 번호", "required": "N", "remarks": ""},
                {"name": "limit", "type": "INTEGER", "description": "페이지당 항목 수", "required": "N", "remarks": ""},
                {"name": "start_date", "type": "STRING", "description": "조회 시작일", "required": "N", "remarks": ""},
                {"name": "end_date", "type": "STRING", "description": "조회 종료일", "required": "N", "remarks": ""},
                {"name": "pharmacy_code", "type": "STRING", "description": "약국 코드", "required": "N", "remarks": ""}
            ],
            "output_params": [
                {"name": "success", "type": "BOOLEAN", "description": "성공 여부", "required": "Y", "remarks": ""},
                {"name": "data", "type": "ARRAY", "description": "직매 매출 목록", "required": "Y", "remarks": ""},
                {"name": "data[].id", "type": "BIGINT", "description": "매출 ID", "required": "Y", "remarks": ""},
                {"name": "data[].pharmacy_code", "type": "STRING", "description": "약국 코드", "required": "N", "remarks": ""},
                {"name": "data[].pharmacy_name", "type": "STRING", "description": "약국명", "required": "N", "remarks": ""},
                {"name": "data[].business_registration_number", "type": "STRING", "description": "사업자등록번호", "required": "Y", "remarks": ""},
                {"name": "data[].address", "type": "STRING", "description": "주소", "required": "N", "remarks": ""},
                {"name": "data[].standard_code", "type": "STRING", "description": "표준 코드", "required": "Y", "remarks": ""},
                {"name": "data[].product_name", "type": "STRING", "description": "제품명", "required": "N", "remarks": ""},
                {"name": "data[].sales_amount", "type": "NUMERIC", "description": "매출액", "required": "N", "remarks": ""},
                {"name": "data[].sales_date", "type": "DATE", "description": "매출일", "required": "N", "remarks": ""},
                {"name": "data[].created_by", "type": "STRING", "description": "생성자", "required": "N", "remarks": ""},
                {"name": "data[].updated_by", "type": "STRING", "description": "수정자", "required": "N", "remarks": ""},
                {"name": "data[].created_at", "type": "TIMESTAMP", "description": "생성일시", "required": "Y", "remarks": ""},
                {"name": "data[].updated_at", "type": "TIMESTAMP", "description": "수정일시", "required": "Y", "remarks": ""},
                {"name": "count", "type": "INTEGER", "description": "전체 항목 수", "required": "Y", "remarks": ""},
                {"name": "page", "type": "INTEGER", "description": "현재 페이지", "required": "Y", "remarks": ""},
                {"name": "limit", "type": "INTEGER", "description": "페이지당 항목 수", "required": "Y", "remarks": ""}
            ]
        },
        {
            "name": "17_실적정보_목록조회",
            "uri": "/performance-records",
            "description": "실적 정보 목록 조회",
            "frequency": "실시간",
            "method": "GET",
            "auth": "Bearer Token",
            "input_params": [
                {"name": "page", "type": "INTEGER", "description": "페이지 번호", "required": "N", "remarks": ""},
                {"name": "limit", "type": "INTEGER", "description": "페이지당 항목 수", "required": "N", "remarks": ""},
                {"name": "settlement_month", "type": "STRING", "description": "정산월", "required": "N", "remarks": ""},
                {"name": "company_id", "type": "STRING", "description": "회사 ID", "required": "N", "remarks": ""},
                {"name": "client_id", "type": "STRING", "description": "고객 ID", "required": "N", "remarks": ""}
            ],
            "output_params": [
                {"name": "success", "type": "BOOLEAN", "description": "성공 여부", "required": "Y", "remarks": ""},
                {"name": "data", "type": "ARRAY", "description": "실적 정보 목록", "required": "Y", "remarks": ""},
                {"name": "data[].id", "type": "BIGINT", "description": "실적 ID", "required": "Y", "remarks": ""},
                {"name": "data[].company_id", "type": "UUID", "description": "회사 ID", "required": "Y", "remarks": ""},
                {"name": "data[].settlement_month", "type": "STRING", "description": "정산월", "required": "Y", "remarks": ""},
                {"name": "data[].prescription_month", "type": "STRING", "description": "처방월", "required": "Y", "remarks": ""},
                {"name": "data[].client_id", "type": "BIGINT", "description": "고객 ID", "required": "Y", "remarks": ""},
                {"name": "data[].product_id", "type": "UUID", "description": "제품 ID", "required": "Y", "remarks": ""},
                {"name": "data[].prescription_qty", "type": "NUMERIC", "description": "처방 수량", "required": "Y", "remarks": ""},
                {"name": "data[].prescription_type", "type": "STRING", "description": "처방 타입", "required": "Y", "remarks": "EDI/직매"},
                {"name": "data[].remarks", "type": "STRING", "description": "비고", "required": "N", "remarks": ""},
                {"name": "data[].registered_by", "type": "UUID", "description": "등록자", "required": "Y", "remarks": ""},
                {"name": "data[].review_status", "type": "STRING", "description": "검토 상태", "required": "Y", "remarks": "대기/승인/반려"},
                {"name": "data[].review_action", "type": "STRING", "description": "검토 액션", "required": "N", "remarks": ""},
                {"name": "data[].updated_by", "type": "UUID", "description": "수정자", "required": "N", "remarks": ""},
                {"name": "data[].commission_rate", "type": "NUMERIC", "description": "수수료율", "required": "N", "remarks": ""},
                {"name": "data[].created_at", "type": "TIMESTAMP", "description": "생성일시", "required": "Y", "remarks": ""},
                {"name": "data[].updated_at", "type": "TIMESTAMP", "description": "수정일시", "required": "Y", "remarks": ""},
                {"name": "count", "type": "INTEGER", "description": "전체 항목 수", "required": "Y", "remarks": ""},
                {"name": "page", "type": "INTEGER", "description": "현재 페이지", "required": "Y", "remarks": ""},
                {"name": "limit", "type": "INTEGER", "description": "페이지당 항목 수", "required": "Y", "remarks": ""}
            ]
        },
        {
            "name": "18_실적흡수율_정보",
            "uri": "/performance-records-absorption",
            "description": "실적 흡수율 분석 정보 조회",
            "frequency": "실시간",
            "method": "GET",
            "auth": "Bearer Token",
            "input_params": [
                {"name": "page", "type": "INTEGER", "description": "페이지 번호", "required": "N", "remarks": ""},
                {"name": "limit", "type": "INTEGER", "description": "페이지당 항목 수", "required": "N", "remarks": ""},
                {"name": "settlement_month", "type": "STRING", "description": "정산월", "required": "N", "remarks": ""},
                {"name": "company_id", "type": "STRING", "description": "회사 ID", "required": "N", "remarks": ""},
                {"name": "client_id", "type": "STRING", "description": "고객 ID", "required": "N", "remarks": ""}
            ],
            "output_params": [
                {"name": "success", "type": "BOOLEAN", "description": "성공 여부", "required": "Y", "remarks": ""},
                {"name": "data", "type": "ARRAY", "description": "흡수율 분석 목록", "required": "Y", "remarks": ""},
                {"name": "count", "type": "INTEGER", "description": "전체 항목 수", "required": "Y", "remarks": ""},
                {"name": "page", "type": "INTEGER", "description": "현재 페이지", "required": "Y", "remarks": ""},
                {"name": "limit", "type": "INTEGER", "description": "페이지당 항목 수", "required": "Y", "remarks": ""}
            ]
        },
        {
            "name": "19_실적증빙파일",
            "uri": "/performance-evidence-files",
            "description": "실적 증빙 파일 정보 조회",
            "frequency": "실시간",
            "method": "GET",
            "auth": "Bearer Token",
            "input_params": [
                {"name": "page", "type": "INTEGER", "description": "페이지 번호", "required": "N", "remarks": ""},
                {"name": "limit", "type": "INTEGER", "description": "페이지당 항목 수", "required": "N", "remarks": ""},
                {"name": "settlement_month", "type": "STRING", "description": "정산월", "required": "N", "remarks": ""},
                {"name": "company_id", "type": "STRING", "description": "회사 ID", "required": "N", "remarks": ""},
                {"name": "client_id", "type": "STRING", "description": "고객 ID", "required": "N", "remarks": ""}
            ],
            "output_params": [
                {"name": "success", "type": "BOOLEAN", "description": "성공 여부", "required": "Y", "remarks": ""},
                {"name": "data", "type": "ARRAY", "description": "증빙 파일 목록", "required": "Y", "remarks": ""},
                {"name": "count", "type": "INTEGER", "description": "전체 항목 수", "required": "Y", "remarks": ""},
                {"name": "page", "type": "INTEGER", "description": "현재 페이지", "required": "Y", "remarks": ""},
                {"name": "limit", "type": "INTEGER", "description": "페이지당 항목 수", "required": "Y", "remarks": ""}
            ]
        },
        {
            "name": "20_정산월_목록조회",
            "uri": "/settlement-months",
            "description": "정산월 목록 조회",
            "frequency": "실시간",
            "method": "GET",
            "auth": "Bearer Token",
            "input_params": [
                {"name": "page", "type": "INTEGER", "description": "페이지 번호", "required": "N", "remarks": ""},
                {"name": "limit", "type": "INTEGER", "description": "페이지당 항목 수", "required": "N", "remarks": ""},
                {"name": "status", "type": "STRING", "description": "상태 필터", "required": "N", "remarks": ""}
            ],
            "output_params": [
                {"name": "success", "type": "BOOLEAN", "description": "성공 여부", "required": "Y", "remarks": ""},
                {"name": "data", "type": "ARRAY", "description": "정산월 목록", "required": "Y", "remarks": ""},
                {"name": "data[].id", "type": "BIGINT", "description": "정산월 ID", "required": "Y", "remarks": ""},
                {"name": "data[].settlement_month", "type": "STRING", "description": "정산월", "required": "Y", "remarks": "YYYY-MM 형식"},
                {"name": "data[].start_date", "type": "DATE", "description": "시작일", "required": "Y", "remarks": ""},
                {"name": "data[].end_date", "type": "DATE", "description": "종료일", "required": "Y", "remarks": ""},
                {"name": "data[].notice", "type": "STRING", "description": "공지사항", "required": "N", "remarks": ""},
                {"name": "data[].status", "type": "STRING", "description": "상태", "required": "Y", "remarks": "active/inactive"},
                {"name": "data[].remarks", "type": "STRING", "description": "비고", "required": "N", "remarks": ""},
                {"name": "data[].created_at", "type": "TIMESTAMP", "description": "생성일시", "required": "Y", "remarks": ""},
                {"name": "count", "type": "INTEGER", "description": "전체 항목 수", "required": "Y", "remarks": ""},
                {"name": "page", "type": "INTEGER", "description": "현재 페이지", "required": "Y", "remarks": ""},
                {"name": "limit", "type": "INTEGER", "description": "페이지당 항목 수", "required": "Y", "remarks": ""}
            ]
        },
        {
            "name": "21_정산내역서_목록조회",
            "uri": "/settlement-share",
            "description": "정산내역서 목록 조회",
            "frequency": "실시간",
            "method": "GET",
            "auth": "Bearer Token",
            "input_params": [
                {"name": "page", "type": "INTEGER", "description": "페이지 번호", "required": "N", "remarks": ""},
                {"name": "limit", "type": "INTEGER", "description": "페이지당 항목 수", "required": "N", "remarks": ""},
                {"name": "settlement_month", "type": "STRING", "description": "정산월", "required": "N", "remarks": ""},
                {"name": "company_id", "type": "STRING", "description": "회사 ID", "required": "N", "remarks": ""}
            ],
            "output_params": [
                {"name": "success", "type": "BOOLEAN", "description": "성공 여부", "required": "Y", "remarks": ""},
                {"name": "data", "type": "ARRAY", "description": "정산내역서 목록", "required": "Y", "remarks": ""},
                {"name": "data[].id", "type": "BIGINT", "description": "정산 ID", "required": "Y", "remarks": ""},
                {"name": "data[].settlement_month", "type": "STRING", "description": "정산월", "required": "Y", "remarks": ""},
                {"name": "data[].company_id", "type": "UUID", "description": "회사 ID", "required": "Y", "remarks": ""},
                {"name": "data[].share_enabled", "type": "BOOLEAN", "description": "공유 활성화 여부", "required": "Y", "remarks": ""},
                {"name": "data[].created_at", "type": "TIMESTAMP", "description": "생성일시", "required": "Y", "remarks": ""},
                {"name": "count", "type": "INTEGER", "description": "전체 항목 수", "required": "Y", "remarks": ""},
                {"name": "page", "type": "INTEGER", "description": "현재 페이지", "required": "Y", "remarks": ""},
                {"name": "limit", "type": "INTEGER", "description": "페이지당 항목 수", "required": "Y", "remarks": ""}
            ]
        }
    ]
    
    # API 폴더 생성
    api_folder = "API_Files"
    if not os.path.exists(api_folder):
        os.makedirs(api_folder)
        print(f"📁 폴더 생성: {api_folder}")
    
    # 각 API 파일 생성
    for api in apis:
        filename = f"{api_folder}/{api['name']}.xlsx"
        create_api_file(api, filename)
    
    print(f"\n🎉 총 {len(apis)}개의 API 파일이 생성되었습니다!")
    print(f"📂 위치: {api_folder}/")

if __name__ == "__main__":
    create_all_api_files()
