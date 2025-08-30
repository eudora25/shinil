#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
from openpyxl import load_workbook
import requests
import json

def check_20_file():
    """20번 파일을 확인합니다."""
    
    try:
        print("=== 20번 파일 (정산월_목록조회) 확인 ===")
        
        # 먼저 로그인해서 토큰을 받아옵니다
        print("로그인하여 토큰 받는 중...")
        login_data = {
            "email": "admin@admin.com",
            "password": "asdf1234"
        }
        
        login_response = requests.post('https://shinil.vercel.app/api/auth', json=login_data)
        
        if login_response.status_code != 200:
            print(f"로그인 실패: {login_response.status_code}")
            return
        
        login_result = login_response.json()
        token = login_result.get('data', {}).get('token')
        
        if not token:
            print("토큰을 받을 수 없습니다.")
            return
        
        print("토큰을 성공적으로 받았습니다.")
        
        # API 요청
        print("API 요청 중...")
        headers = {'Authorization': f'Bearer {token}'}
        response = requests.get('https://shinil.vercel.app/api/settlement-months', headers=headers)
        
        if response.status_code != 200:
            print(f"API 요청 실패: {response.status_code}")
            print(f"응답 내용: {response.text}")
            return
        
        actual_response = response.json()
        print("API 응답을 성공적으로 받았습니다.")
        
        # 실제 응답에서 필드 추출
        actual_fields = []
        for key, value in actual_response.items():
            if key == "data" and isinstance(value, list) and len(value) > 0:
                first_item = value[0]
                for sub_key in first_item.keys():
                    actual_fields.append(f"data[].{sub_key}")
            elif key == "data" and isinstance(value, dict):
                for sub_key in value.keys():
                    actual_fields.append(f"data.{sub_key}")
            else:
                actual_fields.append(key)
        
        print(f"\n실제 API 응답 필드들:")
        for field in actual_fields:
            print(f"  - {field}")
        
        # 엑셀 파일에서 필드 추출
        wb = load_workbook('API_Files/20_정산월_목록조회.xlsx')
        ws = wb.active
        
        # 출력 파라미터 섹션 찾기
        output_section_row = None
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and "출력 파라미터" in str(cell_value):
                output_section_row = row
                break
        
        if not output_section_row:
            print("출력 파라미터 섹션을 찾을 수 없습니다.")
            return
        
        # 출력 파라미터 섹션의 끝 찾기
        output_end_row = output_section_row
        for row in range(output_section_row + 1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and ("입력 파라미터" in str(cell_value) or "응답 예시" in str(cell_value)):
                output_end_row = row - 1
                break
            if row == ws.max_row:
                output_end_row = row
        
        # 엑셀 파일의 필드 추출
        excel_fields = []
        for row in range(output_section_row + 2, output_end_row + 1):  # 헤더 다음 행부터
            field_name = ws.cell(row=row, column=1).value
            if field_name and field_name.strip():
                excel_fields.append(field_name.strip())
        
        print(f"\n엑셀 파일의 필드들:")
        for field in excel_fields:
            print(f"  - {field}")
        
        # 비교
        print(f"\n=== 비교 결과 ===")
        
        # 엑셀에 있지만 실제 API에는 없는 필드들
        missing_in_api = [field for field in excel_fields if field not in actual_fields]
        if missing_in_api:
            print(f"❌ 엑셀에 있지만 API에는 없는 필드들 ({len(missing_in_api)}개):")
            for field in missing_in_api:
                print(f"  - {field}")
        else:
            print("✅ 엑셀의 모든 필드가 API에 존재합니다.")
        
        # API에 있지만 엑셀에는 없는 필드들
        missing_in_excel = [field for field in actual_fields if field not in excel_fields]
        if missing_in_excel:
            print(f"❌ API에 있지만 엑셀에는 없는 필드들 ({len(missing_in_excel)}개):")
            for field in missing_in_excel:
                print(f"  - {field}")
        else:
            print("✅ API의 모든 필드가 엑셀에 존재합니다.")
        
        # 일치하는 필드들
        matching_fields = [field for field in excel_fields if field in actual_fields]
        print(f"✅ 일치하는 필드들 ({len(matching_fields)}개): {', '.join(matching_fields)}")
        
        if not missing_in_api and not missing_in_excel:
            print("\n🎉 완벽하게 일치합니다!")
        else:
            print(f"\n📊 요약:")
            print(f"  - 총 필드 수: API {len(actual_fields)}개, 엑셀 {len(excel_fields)}개")
            print(f"  - 일치: {len(matching_fields)}개")
            print(f"  - 불일치: {len(missing_in_api) + len(missing_in_excel)}개")
        
    except Exception as e:
        print(f"오류 발생: {str(e)}")

if __name__ == "__main__":
    check_20_file()
