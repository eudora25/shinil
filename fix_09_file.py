#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

def fix_09_file():
    """09번 파일을 실제 API 응답 구조에 맞게 수정합니다."""
    
    try:
        # 엑셀 파일 로드
        wb = load_workbook('API_Files/09_공지사항_조회.xlsx')
        ws = wb.active
        
        print("=== 09번 파일을 실제 API 응답 구조로 수정 시작 ===")
        
        # 출력 파라미터 섹션 찾기
        output_section_row = None
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and "출력 파라미터" in str(cell_value):
                output_section_row = row
                print(f"출력 파라미터 섹션 발견: 행 {row}")
                break
        
        if output_section_row:
            # 출력 파라미터 섹션의 끝 찾기
            output_end_row = output_section_row
            for row in range(output_section_row + 1, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=1).value
                if cell_value and any(keyword in str(cell_value) for keyword in ["입력 파라미터", "응답 예시"]):
                    break
                output_end_row = row
            
            print(f"출력 파라미터 섹션 범위: 행 {output_section_row} ~ {output_end_row}")
            
            # 기존 출력 파라미터 행들 삭제 (헤더 제외)
            rows_to_delete = list(range(output_section_row + 2, output_end_row + 1))
            for row_num in reversed(rows_to_delete):
                ws.delete_rows(row_num)
            
            print("기존 출력 파라미터 행들 삭제 완료")
            
            # 실제 API 응답 구조에 맞는 새로운 출력 파라미터 데이터 정의
            new_output_params = [
                ["success", "BOOLEAN", "성공 여부", "Y", "true/false"],
                ["message", "STRING", "응답 메시지", "Y", "공지사항 목록 조회 결과 메시지"],
                ["data[].id", "STRING", "공지사항 ID", "Y", "고유 식별자"],
                ["data[].title", "STRING", "제목", "Y", "공지사항 제목"],
                ["data[].content", "STRING", "내용", "Y", "공지사항 내용"],
                ["data[].is_pinned", "BOOLEAN", "고정 여부", "Y", "상단 고정 여부"],
                ["data[].view_count", "INTEGER", "조회수", "Y", "조회 횟수"],
                ["data[].created_by", "STRING", "작성자", "Y", "작성한 사용자 ID"],
                ["data[].created_at", "STRING", "작성일시", "Y", "ISO 8601 형식"],
                ["data[].updated_at", "STRING", "수정일시", "N", "ISO 8601 형식"],
                ["data[].file_url", "STRING", "첨부파일 URL", "N", "첨부파일 URL 배열"],
                ["data[].links", "STRING", "링크", "N", "관련 링크"],
                ["data[].updated_by", "STRING", "수정자", "N", "수정한 사용자 ID"]
            ]
            
            # 새로운 출력 파라미터 추가
            current_row = output_section_row + 2  # 헤더 다음 행부터 시작
            
            for param in new_output_params:
                ws.cell(row=current_row, column=1, value=param[0])  # 파라미터명
                ws.cell(row=current_row, column=2, value=param[1])  # 타입
                ws.cell(row=current_row, column=3, value=param[2])  # 설명
                ws.cell(row=current_row, column=4, value=param[3])  # 필수
                ws.cell(row=current_row, column=5, value=param[4])  # 비고
                
                # 스타일 적용
                for col in range(1, 6):
                    cell = ws.cell(row=current_row, column=col)
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                
                current_row += 1
            
            print("새로운 출력 파라미터 추가 완료")
            
            # 파일 저장
            wb.save('API_Files/09_공지사항_조회.xlsx')
            print("✅ 09번 파일을 실제 API 응답 구조로 수정 완료!")
            
            # 수정된 내용 확인
            print("\n=== 수정된 출력 파라미터 ===")
            for row in range(output_section_row, current_row):
                row_data = []
                for col in range(1, 6):
                    cell_value = ws.cell(row=row, column=col).value
                    row_data.append(str(cell_value) if cell_value is not None else "")
                
                if any(row_data):
                    print(f"  행 {row}: {' | '.join(row_data)}")
        
    except Exception as e:
        print(f"오류가 발생했습니다: {str(e)}")

if __name__ == "__main__":
    fix_09_file()
